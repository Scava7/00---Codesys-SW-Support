import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox

def scegli_file(titolo, tipi_file):
    file_path = filedialog.askopenfilename(title=titolo, filetypes=tipi_file)
    return file_path

def scegli_salvataggio(titolo, tipo_file):
    file_path = filedialog.asksaveasfilename(title=titolo, defaultextension=tipo_file[1], filetypes=[tipo_file])
    return file_path

def esporta_sqlite_in_excel(db_path):
    xlsx_path = scegli_salvataggio("Salva il file Excel", ("Excel Workbook", ".xlsx"))
    if not xlsx_path:
        return

    conn = sqlite3.connect(db_path)
    xls_writer = pd.ExcelWriter(xlsx_path, engine="openpyxl")

    query = "SELECT name FROM sqlite_master WHERE type='table';"
    tables = pd.read_sql(query, conn)

    for table_name in tables['name']:
        df = pd.read_sql(f"SELECT * FROM '{table_name}'", conn)

        if table_name == "PAR_BOOL" and "factory_value" in df.columns:
            df["factory_value"] = df["factory_value"].apply(lambda x: "TRUE" if str(x).strip().upper() == "TRUE" else "FALSE")

        df.to_excel(xls_writer, sheet_name=table_name, index=False)

    xls_writer.close()
    conn.close()

    wb = load_workbook(xlsx_path)
    for table_name in tables['name']:
        ws = wb[table_name]
        headers = [cell.value for cell in ws[1]]
        if "group" in headers:
            group_index = headers.index("group") + 1
            group_colors = {}
            pastel_colors = [
                "FFCCCC", "FFE5CC", "FFFFCC", "E5FFCC", "CCFFE5",
                "CCFFFF", "CCE5FF", "CCCCFF", "E5CCFF", "FFCCFF",
            ]

            def get_color(value):
                idx = abs(hash(value)) % len(pastel_colors)
                color = pastel_colors[idx]
                return PatternFill(start_color=color, end_color=color, fill_type="solid")

            for row in ws.iter_rows(min_row=2):
                group_val = row[group_index - 1].value
                if group_val not in group_colors:
                    group_colors[group_val] = get_color(group_val)
                fill = group_colors[group_val]
                for cell in row:
                    cell.fill = fill

    wb.save(xlsx_path)
    messagebox.showinfo("Completato", f"Esportazione completata in:\n{xlsx_path}")
